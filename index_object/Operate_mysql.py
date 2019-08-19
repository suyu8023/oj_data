'''
2019/08/15
数据库：ojdata
数据表：info,contests,rank_log
info : 存放用户账号密码等信息，以及主页显示总排名信息
contests : 存放每一场比赛详情
rank_log : 用于显示最近每场比赛积分情况

permission权限：student,teacher,admin
add_judge:１个人，２已添加，３团队赛未添加

'''
import pymysql
import openpyxl
import datetime
import index_object.Data_Formatting
from dateutil.relativedelta import relativedelta

class Operate_mysql():

    # 创建数据表
    # username ,name ,password ,grade ,school ,oj ,vj_account,vj,nc_account,nc,cf_account,cf_count,cf ,rank,permission
    # create table if not exists info(username varchar(15),name varchar(10),password varchar(10),grade varchar(5),school varchar(10),oj float,vj_account varchar(15),vj float,nc_account varchar(15),nc float,cf_account varchar(15),cf float,rank float,permission varchar(10))
    # create table if not exists contests(username varchar(15),ojclass varchar(5),cid varchar(10),cid_time TIMESTAMP,pid varchar(5),ac_time varchar(20),submissions int,difficult_weight float,time_weight float)
    # create table if not exists rank_log(username varchar(15),ojclass varchar(10),cid varchar(10),cid_time TIMESTAMP,intergration float,judge_add varchar(2))

    Format = index_object.Data_Formatting.Data_Formatting

    # 总排名
    def show_rank(self):
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        sql = 'select school,grade,username,name,oj,vj,nc,cf,rank from info order by rank'
        cur.execute(sql)
        cols = cur.fetchall()
        con.close()
        if cols.count() != '':
            return self.Format.show_rank(cols)
        return False

    # 条件查询
    def select_public(self, username,name,grade,school):
        user ,nam,gra,sch = '','','',''
        if username != '':
            user = 'username like "%' + username + '%"'
        if name != '':
            nam = 'name like "%' + name + '%"'
        if grade != '':
            gra = 'grade like "%' + grade + '%"'
        if school != '':
            sch = 'school like "%' + school + '%"'
        join_str = ' and '
        join_lis = [user if user != '', nam if nam != '', gra if gra != '', sch if sch != '']
        select_where = join_str.join(join_lis)

        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        sql = 'select school,grade,username,name,oj,vj,nc,cf,rank from contests where' + select_where
        cur.execute(sql)
        cols = cur.fetchall()
        con.close()
        if cols != '':
            return self.Format.show_rank(cols)
        return False

    #　进入个人比赛页面
    def select_person(self, username,start_time,end_time):
        if end_time == '':
            end_time = datetime.datetime.now()
        if start_time == '':
            start_time = end_time - relativedelta(months=+1)
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        sql = 'select username,ojclass,cid,intergration from rank_log where username = %s and cid_time > %d and cid_time < %d' % \
              ('"' + username + '"',start_time,end_time)
        cur.execute(sql)
        cols = cur.fetchall()
        con.close()
        if cols.count() != '':
            return self.Format.select_private(cols)
        return False

    # 查看单场比赛情况
    def select_contests(self,username,cid):
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        sql = 'select username,ojclass,cid,cid_time,pid,ac_time.submissions,difficult_weight,time_weight from contests where username = %s and cid = %s' % \
              ('"' + username + '"', '"' + str(cid) + '"')
        cur.execute(sql)
        cols = cur.fetchall()
        con.close()
        if cols.count() != '':
            return self.Format.select_private(cols)
        return False

    # 批量添加人员
    def add_persons(self,filename):
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()

        wb = openpyxl.load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[0]]
        for i in range(sheet.max_row):
            username = sheet.cell(row = i+1,column=1).value
            name = sheet.cell(row = i+1, column=2).value
            password = sheet.cell(row = i+1, column=3).value
            grade = sheet.cell(row = i+1, column=4).value
            school = sheet.cell(row = i+1, column=4).value
            sql = "insert into info values ('%s','%s','%s','%s','%s','%f','%s','%f','%s','%f','%s','%f','%f','%s')" % (
                username, name, password, grade, school, 0, '', 0, '', 0, '', 0, 0,'student')
            cur.execute(sql)
            con.commit()
        con.close()

    # 添加个人
    def add_person(self,username,name,password,grade,school):
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        try:
            sql = "insert into info values ('%s','%s','%s','%s','%s','%f','%s','%f','%s','%f','%s','%f','%f','%s')" % (
                username, name, password, grade, school, 0,'',0,'',0,'',0,0,'student')
            cur.execute(sql)
            con.commit()
            con.close()
        except:
            return False

    #修改信息
    def update_message(self,username,password):
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        try:
            sql = "update info set password = '%s' from info where username = '%s'" % (
                '"' + password + '"', '"' + username + '"')
            cur.execute(sql)
            con.commit()
            con.close()
        except:
            return False

    # 登录判断
    def judge_login(self,username, password):
        con = pymysql.connect("localhost", "root", "acm506", "ojdata")
        cur = con.cursor()
        try:
            sql = "select name,permission from info where password = '%s' and username = '%s'" % (
                '"' + password + '"', '"' + username + '"')
            cur.execute(sql)
            cols = cur.fetchall()
            con.close()
            if cols.count() == 0:
                return False
            else:
                return cols[14]
        except:
            return False

    # 团队赛，成绩附录
    def results_dubbing(self,username,cid,username1,username2,username3):

